#!/usr/bin/python3
# -*- coding: utf-8 -*-

from io import BytesIO
from collections import OrderedDict
from typing import Any, List, Dict, Generator, Union
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

class Excel:

    def __init__(self, data: Dict[str, Union[str, List[Any]]]) -> None:
        self.json_dict: Dict[str, Union[str, List[Any]]] = data
        self.bytes: Union[BytesIO, None] = None
        self.size: int = 0

    def get_dict_attribute_list(
        self, dict_list: Generator[Dict[str, Union[str, List[Any]]], None, None]) -> None:
        '''
        遍历所有字典, 获取字典所有属性名组成表头(保存属性顺序, 去除重复), 保存字典列表

        :param dict_list list: 字典列表
        :returns: None
        '''
        self.excel_list: List[Dict[str, str]] = []
        temp_dict: Dict[str, str] = OrderedDict()
        for item in dict_list:
            self.excel_list.append(item)
            temp_dict |= item

        self.header: List[str] = list(temp_dict)

    def flatten_refactor(self, dic: Dict[str, Union[str, List[Any]]]
        ) -> Generator[Dict[str, Union[str, List[Any]]], None, None]:
        '''
        逐一返回最底层的字典数据(无 children 属性)

        :param dic dict: 传入一个多层字典,
        :returns: 生成器
        '''
        try:
            for child in dic['children']:
                for k in [ 'id', 'label' ]: child.pop(k, None)
                yield from self.flatten_refactor(child)
        except Exception:
            for k in [ 'id', 'label' ]: dic.pop(k, None)
            yield dic

    def create_excel_style(self) -> None:
        '''
        定义字体, 边框, 对齐, 颜色样式
        '''
        side = Side(border_style='medium', color='000000')
        self.border = Border(*([ side ] * 4))
        self.alignment = Alignment(
            horizontal    = 'center',
            vertical      = 'center',
            text_rotation = 0,
            wrap_text     = False,
            shrink_to_fit = False,
            indent        = 0
        )
        self.header_font = Font(name='Microsoft JhengHei', color='FFFFFF', size=10, bold=True)
        self.header_patternfill = PatternFill('solid', fgColor='366092')

    def set_hyperlink(self, link: str, title: str) -> str:
        '''
        设置超链接(未修改链接样式)

        :param link str: 链接
        :param title str: 链接显示内容
        :return str: excel 超链接写法字符串
        '''
        return f'''=HYPERLINK("{link}","{title}")'''

    def set_header_style(self) -> None:
        '''
        设置表头样式 width 为第一栏表头对应的宽度
        '''
        self.sheet.row_dimensions[1].height = 20
        headers = list(map(str.capitalize, self.header))
        headers[self.header.index('title')] = 'Name'
        headers[self.header.index('value')] = 'Path'
        self.sheet.append(headers)
        width = { 2: 8, 3: 10, 5: 13, 6: 12, 7: 11, 8: 8, 10: 45 }
        for i in range(len(self.sheet[1])):
            cell, letter = self.sheet[1][i], chr(65 + i)
            self.sheet.column_dimensions[letter].width = width.get(i, 20)
            cell.font = self.header_font
            cell.border = self.border
            cell.fill = self.header_patternfill
            cell.alignment = self.alignment

    def create_excel(self) -> None:
        '''
        创建 excel 表, 添加样式, 保存成流(stream) 供传输
        '''
        wb: Workbook = Workbook()
        self.sheet = wb.active
        self.sheet.title = 'SMS 工具清单'
        self.set_header_style()

        url_index: int = len(self.header)
        for i, dic in enumerate(self.excel_list):

            dic['link'] = self.set_hyperlink(dic['link'], dic['title'])
            self.sheet.append([ str(dic.get(key, '')) for key in self.header ])
            self.sheet.cell(row=i + 2, column=url_index).style = 'Hyperlink'

            for cell in self.sheet[ i + 2 ]:
                cell.border = self.border

        self.bytes = BytesIO()
        wb.save(self.bytes)
        self.bytes.seek(0)
        wb.close()
        self.size = self.bytes.getbuffer().nbytes

    def main(self) -> None:
        dict_list = self.flatten_refactor(self.json_dict)
        self.get_dict_attribute_list(dict_list)
        self.create_excel_style()
        self.create_excel()
